from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

# Imorting the necessary modules
try:
        from openpyxl.cell import get_column_letter
except ImportError:
        from openpyxl.utils import get_column_letter
        from openpyxl.utils import column_index_from_string

import pandas as pd
import datetime
import time
import os
from pathlib import Path, PureWindowsPath

def get_current_time():
    return datetime.datetime.fromtimestamp(time.time()).strftime('%Y-%m-%d %H:%M:%S')

def make_dataframe():

    Cream_temperature = str(input("What is the temperature of the cream mix? "))
    Ice_temperature = str(input("What is the temperature of the ice bath? "))

    try:
        Cream_temperature = int(Cream_temperature)
        Ice_temperature = int(Ice_temperature)
    except ValueError:
        Cream_temperature = float(Cream_temperature)
        Ice_temperature = float(Ice_temperature)

    Cream_Temp_list = []
    Ice_Temp_list = []
    time_list = []


    Cream_Temp_list.append(f'{Cream_temperature}°C')
    Ice_Temp_list.append(f'{Ice_temperature}°C')
    time_list.append(get_current_time())

    Continue = str(input("Do you want to add more data? "))

    while Continue.lower() == "yes" or Continue.lower() == "y":
        Cream_temperature =str(input("What is the temperature of the cream mix? "))
        Ice_temperature = str(input("What is the temperature of the ice bath? "))

        try:
            Cream_temperature = int(Cream_temperature)
            Ice_temperature = int(Ice_temperature)
        except ValueError:
            Cream_temperature = float(Cream_temperature)
            Ice_temperature = float(Ice_temperature)

        Cream_Temp_list.append(f'{Cream_temperature}°C')
        Ice_Temp_list.append(f'{Ice_temperature}°C')
        time_list.append(get_current_time())

        Continue = str(input("Do you want to add more data? "))

        if Continue.lower() == "no" or Continue.lower() == "n":
            break

    data = {
    'Time': time_list,
    'Temperature of Cream Mix': Cream_Temp_list,
    'Temperature of Ice Bath': Ice_Temp_list
    }

    df = pd.DataFrame(data)

    print(f"\n{df}")

    return df

FileName = "IceCream Measurements"

def fit_content():
    workbook = load_workbook(f"{FileName}.xlsx")
    for sheet_name in workbook.sheetnames:
        for column_cells in workbook[sheet_name].columns:
                new_column_length = max(len(str(cell.value)) for cell in column_cells)
                new_column_letter = (get_column_letter(column_cells[0].column))
                if new_column_length > 0:
                    workbook[sheet_name].column_dimensions[new_column_letter].width = new_column_length*1.23
        workbook.save(f"{FileName}.xlsx")

def create_file(FileName):
    # data = {
    # 'Time': [''],
    # 'Temperature of Cream Mix': [''],
    # 'Temperature of Ice Bath': ['']
    # }

    # df = pd.DataFrame(data)

    workbook = Workbook()
    workbook.save(f"{FileName}.xlsx")

    # openworkbook = load_workbook(f"{FileName}.xlsx")
    # openSheet = openworkbook.active

    # for row in dataframe_to_rows(df, index=False, header=True):
    #     openSheet.append(row)

    # openworkbook.save(f"{FileName}.xlsx")

def check_file(FileName):
    try:
        create_file(FileName)
    except FileExistsError:
        print("File already exists")
        os.rename(f"{FileName}.xlsx", f"{FileName}.xlsx.old")
        if os.path.exists(f"{FileName}.xlsx.old"):
            os.remove(f"{FileName}.xlsx.old")
        create_file(FileName)
    except FileNotFoundError:
        print("File does not exist")
        create_file(FileName)

def add_dataframe(FileName, df):

    openworkbook = load_workbook(f"{FileName}.xlsx")
    openSheet = openworkbook.active

    for row in dataframe_to_rows(df, index=False, header=True):
        openSheet.append(row)

    openworkbook.save(f"{FileName}.xlsx")


def add_reflexion(FileName):
    reflexion = str(input("\nEnter the reflexion: "))

    print(f"\n{reflexion}\n")

    openworkbook = load_workbook(f"{FileName}.xlsx")
    openSheet = openworkbook.active

    openSheet['E1'] = "Reflexion:"

    openSheet['E2'].alignment = Alignment(horizontal="center",
                                        vertical="center")
    openSheet['E2'].alignment = Alignment(wrap_text=True)
    openSheet.merge_cells('E2:E4')
    openSheet['E2'] = reflexion

    openworkbook.save(f"{FileName}.xlsx")

def place_images(FileName):
    IceTemp1 = Image('imgs\\IceBath\\IceTemp1.jpg')
    IceTemp2 = Image('imgs\\IceBath\\IceTemp2.jpg')
    IceTemp3 = Image('imgs\\IceBath\\IceTemp3.jpg')
    IceTemp4 = Image('imgs\\IceBath\\IceTemp4.jpg')
    IceTemp5 = Image('imgs\\IceBath\\IceTemp5.jpg')

    CreamTemp1 = Image('imgs\\CreamMix\\CreamTemp1.jpg')
    CreamTemp2 = Image('imgs\\CreamMix\\CreamTemp2.jpg')
    CreamTemp3 = Image('imgs\\CreamMix\\CreamTemp3.jpg')
    CreamTemp4 = Image('imgs\\CreamMix\\CreamTemp4.jpg')
    CreamTemp5 = Image('imgs\\CreamMix\\CreamTemp5.jpg')

    openworkbook = load_workbook(f"{FileName}.xlsx")
    openSheet = openworkbook.active

    openSheet.add_image(IceTemp1, 'C9')
    openSheet.add_image(IceTemp2, 'C10')
    openSheet.add_image(IceTemp3, 'C11')
    openSheet.add_image(IceTemp4, 'C12')
    openSheet.add_image(IceTemp5, 'C13')

    openSheet.add_image(CreamTemp1, 'B9')
    openSheet.add_image(CreamTemp2, 'B10')
    openSheet.add_image(CreamTemp3, 'B11')
    openSheet.add_image(CreamTemp4, 'B12')
    openSheet.add_image(CreamTemp5, 'B13')

    openworkbook.save(f"{FileName}.xlsx")

def start():
    check_file(FileName)
    add_dataframe(FileName, make_dataframe())
    add_reflexion(FileName)
    fit_content()

if __name__ == "__main__":
    #add_reflexion(FileName)
    place_images(FileName)
    fit_content()